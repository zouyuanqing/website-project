from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, make_response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash
from werkzeug.datastructures import CombinedMultiDict
import os
import uuid
import json
import csv
import io
import zipfile
import shutil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from urllib.parse import quote
import time
import random

from config import Config
from models import db, User, Admin, Form, FormField, Submission, SubmissionData, UploadFile, PaymentOrder, PaymentAccount
from forms import LoginForm, RegisterForm, AdminLoginForm, CreateFormForm, EditFormForm, DynamicForm, FormFieldForm

def encode_filename_for_http(filename):
    """对文件名进行HTTP头兼容的编码处理"""
    # 使用RFC 5987标准处理非ASCII文件名
    encoded_filename = quote(filename, safe='')
    return f"attachment; filename*=UTF-8''{encoded_filename}"

def generate_order_no(payment_type='PAY'):
    """生成唯一订单号"""
    timestamp = int(time.time() * 1000)  # 毫秒时间戳
    random_num = random.randint(1000, 9999)  # 4位随机数
    return f"{payment_type}{timestamp}{random_num}"

def create_app():
    from payment_config import get_payment_processor, PaymentResult
    from dotenv import load_dotenv

    # 加载环境变量
    load_dotenv()

    # 修复缩进问题
    app = Flask(__name__)
    app.config.from_object(Config)

    # 初始化扩展
    db.init_app(app)

    # 登录管理
    login_manager = LoginManager()
    login_manager.init_app(app)
    login_manager.login_view = 'login'  # type: ignore
    login_manager.login_message = '请先登录'
    # 允许同一账号在多个设备同时登录
    login_manager.session_protection = None  # type: ignore

    @login_manager.user_loader
    def load_user(user_id):
        if user_id.startswith('admin_'):
            admin_id = user_id.replace('admin_', '')
            return Admin.query.get(int(admin_id))
        return User.query.get(int(user_id))

    # 创建上传目录
    upload_dir = os.path.join(app.instance_path, app.config['UPLOAD_FOLDER'])
    os.makedirs(upload_dir, exist_ok=True)
    
    # 添加请求日志中间件
    @app.before_request
    def log_request_info():
        if request.path.startswith('/admin/export/') or request.path.startswith('/admin/test'):
            app.logger.info(f"收到请求: {request.method} {request.path} - 参数: {request.args} - 用户登录状态: {current_user.is_authenticated if current_user else 'None'}")

    def allowed_file(filename):
        return '.' in filename and \
               filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']
        
    def save_uploaded_file(file, field_name, submission_id):
        """保存上传的文件"""
        try:
            app.logger.info(f"📁 处理文件上传: field={field_name}, file={file}, filename={getattr(file, 'filename', 'N/A')}")
            
            if file and allowed_file(file.filename):
                # 检查文件大小
                if file.content_length and file.content_length > app.config['MAX_CONTENT_LENGTH']:
                    raise ValueError(f"文件大小超过限制 {app.config['MAX_CONTENT_LENGTH']/1024/1024:.0f}MB")
                
                # 生成唯一文件名
                file_ext = file.filename.rsplit('.', 1)[1].lower()
                saved_filename = f"{uuid.uuid4().hex}.{file_ext}"
                file_path = os.path.join(upload_dir, saved_filename)
                
                app.logger.info(f"📁 保存文件到: {file_path}")
                
                # 保存文件
                file.save(file_path)
                
                # 检查实际文件大小
                actual_size = os.path.getsize(file_path)
                if actual_size > app.config['MAX_CONTENT_LENGTH']:
                    os.remove(file_path)  # 删除超大文件
                    raise ValueError(f"文件大小超过限制 {app.config['MAX_CONTENT_LENGTH']/1024/1024:.0f}MB")
                
                app.logger.info(f"✅ 文件保存成功: {file.filename} -> {saved_filename}, 大小: {actual_size} bytes")
                
                # 保存文件信息到数据库
                upload_file = UploadFile(
                    submission_id=submission_id,
                    field_name=field_name,
                    original_filename=file.filename,
                    saved_filename=saved_filename,
                    file_size=actual_size,
                    file_type=file.content_type or 'application/octet-stream'
                )
                db.session.add(upload_file)
                return upload_file
            else:
                app.logger.warning(f"⚠️ 文件上传失败: 文件不存在或格式不支持, file={file}, filename={getattr(file, 'filename', 'N/A')}")
        except Exception as e:
            app.logger.error(f"❌ 文件上传失败: {str(e)}")
            flash(f"文件上传失败: {str(e)}", 'danger')
        return None
    
    # 路由定义
    @app.route('/')
    def index():
    return render_template('index.html')
    
    @app.route('/login', methods=['GET', 'POST'])
    def login():
    if current_user.is_authenticated and not current_user.get_id().startswith('admin_'):
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        # 查找用户（通过邮箱或手机号）
        user = User.query.filter(
            (User.email == form.login_id.data) | (User.phone == form.login_id.data)
        ).first()
        
        if user and user.check_password(form.password.data):
            login_user(user, remember=form.remember_me.data)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        flash('用户名或密码错误', 'danger')
    
    return render_template('auth/login.html', form=form)

    @app.route('/register', methods=['GET', 'POST'])
    def register():
    form = RegisterForm()
    if form.validate_on_submit():
        # 检查用户是否已存在
        existing_user = None
        if form.email.data:
            existing_user = User.query.filter_by(email=form.email.data).first()
        if not existing_user and form.phone.data:
            existing_user = User.query.filter_by(phone=form.phone.data).first()
        
        if existing_user:
            flash('该邮箱或手机号已被注册', 'danger')
            return render_template('auth/register.html', form=form)
        
        # 创建新用户
        user = User(
            name=form.name.data,
            email=form.email.data if form.email.data else None,
            phone=form.phone.data if form.phone.data else None
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        
        flash('注册成功，请登录', 'success')
        return redirect(url_for('login'))
    
    return render_template('auth/register.html', form=form)

    @app.route('/logout')
        @login_required
    def logout():
    logout_user()
    return redirect(url_for('index'))

    @app.route('/dashboard')
        @login_required
    def dashboard():
    if current_user.get_id().startswith('admin_'):
        return redirect(url_for('admin_dashboard'))
    
    # 用户仪表板
    submissions = Submission.query.filter_by(user_id=current_user.id).order_by(Submission.submitted_at.desc()).all()
    
    # 用户支付统计数据
    user_payment_stats = {
        'total_payments': 0,
        'paid_orders': 0, 
        'pending_payments': 0,
        'total_paid_amount': 0.0,
        'wechat_payments': 0,
        'alipay_payments': 0,
        'wechat_amount': 0.0,
        'alipay_amount': 0.0
    }
    
    # 获取用户所有支付订单
    user_payment_orders = PaymentOrder.query.join(Submission).filter(
        Submission.user_id == current_user.id
    ).all()
    
    if user_payment_orders:
        user_payment_stats['total_payments'] = len(user_payment_orders)
        user_payment_stats['paid_orders'] = len([p for p in user_payment_orders if p.status == 'paid'])
        user_payment_stats['pending_payments'] = len([p for p in user_payment_orders if p.status == 'pending'])
        user_payment_stats['total_paid_amount'] = sum([p.amount for p in user_payment_orders if p.status == 'paid'])
        
        # 按支付方式统计
        wechat_orders = [p for p in user_payment_orders if p.payment_type == 'wechat_pay']
        alipay_orders = [p for p in user_payment_orders if p.payment_type == 'alipay']
        
        user_payment_stats['wechat_payments'] = len(wechat_orders)
        user_payment_stats['alipay_payments'] = len(alipay_orders)
        user_payment_stats['wechat_amount'] = sum([p.amount for p in wechat_orders if p.status == 'paid'])
        user_payment_stats['alipay_amount'] = sum([p.amount for p in alipay_orders if p.status == 'paid'])
    
    return render_template('user/dashboard.html', 
                         submissions=submissions, 
                         payment_stats=user_payment_stats,
                         payment_orders=user_payment_orders)

    @app.route('/available_forms')
        @login_required
    def available_forms():
    """可用表单页面"""
    if current_user.get_id().startswith('admin_'):
        return redirect(url_for('admin_dashboard'))
    
    # 获取所有活跃的表单
    active_forms = Form.query.filter_by(is_active=True).order_by(Form.updated_at.desc()).all()
    
    # 获取当前用户已提交的表单
    user_submissions = Submission.query.filter_by(user_id=current_user.id).all()
    
    # 获取当前用户已提交的表单ID列表
    user_submitted_form_ids = [submission.form_id for submission in user_submissions]
    
    return render_template('user/available_forms.html', 
                          forms=active_forms,
                          submitted_form_ids=user_submitted_form_ids,
                          submissions=user_submissions)

    @app.route('/user/payment-history')
        @login_required 
    def user_payment_history():
    """用户支付历史页面"""
    if current_user.get_id().startswith('admin_'):
        return redirect(url_for('admin_dashboard'))
    
    # 获取用户所有支付记录，支持分页
    page = request.args.get('page', 1, type=int)
    per_page = 20  # 每页显示20条记录
    
    # 获取状态筛选
    status_filter = request.args.get('status', '')
    payment_type_filter = request.args.get('payment_type', '')
    
    # 构建查询
    query = PaymentOrder.query.join(Submission).filter(
        Submission.user_id == current_user.id
    )
    
    if status_filter:
        query = query.filter(PaymentOrder.status == status_filter)
    if payment_type_filter:
        query = query.filter(PaymentOrder.payment_type == payment_type_filter)
    
    # 执行分页查询
    payment_orders = query.order_by(PaymentOrder.created_at.desc()).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )
    
    # 统计数据
    payment_stats = {
        'total_payments': query.count(),
        'paid_orders': query.filter(PaymentOrder.status == 'paid').count(),
        'pending_payments': query.filter(PaymentOrder.status == 'pending').count(),
        'total_paid_amount': db.session.query(db.func.sum(PaymentOrder.amount)).join(Submission).filter(
            Submission.user_id == current_user.id,
            PaymentOrder.status == 'paid'
        ).scalar() or 0.0
    }
    
    return render_template('user/payment_history.html', 
                         payment_orders=payment_orders,
                         payment_stats=payment_stats,
                         status_filter=status_filter,
                         payment_type_filter=payment_type_filter)
    
    @app.route('/form/<int:form_id>', methods=['GET', 'POST'])
    def view_form(form_id):
        form_obj = Form.query.get_or_404(form_id)
        if not form_obj.is_active:
            flash('该表单已停用', 'warning')
            return redirect(url_for('index'))
        
        # 如果用户未登录，跳转到登录页面
        if not current_user.is_authenticated:
            return redirect(url_for('login', next=request.url))
        
        # 检查用户是否已提交过此表单
        existing_submission = Submission.query.filter_by(
            form_id=form_id,
            user_id=current_user.id
        ).first()
        
        if existing_submission:
            flash('您已提交过此表单', 'info')
            return redirect(url_for('view_submission', submission_id=existing_submission.id))
        
        # 创建动态表单
        form_fields = FormField.query.filter_by(form_id=form_id).order_by(FormField.order_index).all()
        # 修复文件上传问题：需要传递 files 数据
        formdata = None
        if request.method == 'POST':
            # 将 form 和 files 数据合并为 CombinedMultiDict
            formdata = CombinedMultiDict([request.files, request.form])
        
        dynamic_form = DynamicForm(form_fields, formdata=formdata)
        
        if request.method == 'POST' and dynamic_form.validate():
            try:
                # 创建提交记录
                submission = Submission(
                    form_id=form_id,
                    user_id=current_user.id
                )
                db.session.add(submission)
                db.session.flush()  # 获取submission.id
                
                # 保存表单数据
                payment_orders = []  # 存储支付订单
                
                for field in form_fields:
                    form_field = getattr(dynamic_form, field.field_name, None)
                    if form_field:
                        field_value = form_field.data
                        
                        app.logger.info(f"📄 处理字段: {field.field_name} (类型: {field.field_type}), 值: {type(field_value)} - {field_value}")
                        
                        if field.field_type == 'file' and field_value:
                            # 处理文件上传
                            app.logger.info(f"📁 开始处理文件上传: {field.field_name}")
                            if not save_uploaded_file(field_value, field.field_name, submission.id):
                                # 文件上传失败，回滚提交
                                app.logger.error(f"❌ 文件上传失败，回滚事务: {field.field_name}")
                                db.session.rollback()
                                return render_template('user/form.html', form_obj=form_obj, form=dynamic_form)
                        elif field.field_type in ['wechat_pay', 'alipay'] and field_value:
                            # 处理支付字段
                            try:
                                amount = float(field_value)
                                if amount > 0:
                                    # 创建支付订单
                                    order_no = generate_order_no()
                                    payment_order = PaymentOrder(
                                        submission_id=submission.id,
                                        field_name=field.field_name,
                                        payment_type=field.field_type,
                                        amount=amount,
                                        order_no=order_no,
                                        status='pending',
                                        payment_account_id=field.payment_account_id  # 关联收款账户
                                    )
                                    db.session.add(payment_order)
                                    payment_orders.append(payment_order)
                                    
                                    app.logger.info(f"💰 创建支付订单: {order_no}, 金额: {amount}元, 收款账户ID: {field.payment_account_id}")
                                    
                                    # 保存支付金额到提交数据中
                                    submission_data = SubmissionData(
                                        submission_id=submission.id,
                                        field_name=field.field_name,
                                        field_value=str(amount)
                                    )
                                    db.session.add(submission_data)
                            except (ValueError, TypeError) as e:
                                app.logger.error(f"❌ 支付金额无效: {field_value}, 错误: {str(e)}")
                                flash(f'支付金额格式错误: {field.field_label}', 'danger')
                                db.session.rollback()
                                return render_template('user/form.html', form_obj=form_obj, form=dynamic_form)
                        elif field_value is not None and field_value != '':
                            # 处理复选框数据
                            if field.field_type == 'checkbox' and isinstance(field_value, list):
                                field_value = ','.join(field_value)
                            
                            # 保存普通字段数据
                            submission_data = SubmissionData(
                                submission_id=submission.id,
                                field_name=field.field_name,
                                field_value=str(field_value)
                            )
                            db.session.add(submission_data)
                
                db.session.commit()
                
                # 检查是否有支付订单
                if payment_orders:
                    app.logger.info(f"💳 检测到 {len(payment_orders)} 个支付订单，跳转到支付页面")
                    flash('表单提交成功！请完成支付。', 'success')
                    return redirect(url_for('payment_page', submission_id=submission.id))
                else:
                    flash('表单提交成功！', 'success')
                    return redirect(url_for('view_submission', submission_id=submission.id))
                
            except Exception as e:
                db.session.rollback()
                app.logger.error(f"表单提交失败: {str(e)}")
                flash('表单提交失败，请重试', 'danger')
        
        elif request.method == 'POST':
            # 表单验证失败
            flash('请检查表单信息是否正确填写', 'warning')
        
        return render_template('user/form.html', form_obj=form_obj, form=dynamic_form)
    
    @app.route('/submission/<int:submission_id>')
    @login_required
    def view_submission(submission_id):
        submission = Submission.query.get_or_404(submission_id)
        
        # 验证权限
        if not current_user.get_id().startswith('admin_') and submission.user_id != current_user.id:
            flash('无权访问', 'danger')
            return redirect(url_for('dashboard'))
        
        return render_template('user/submission.html', submission=submission)
    
    @app.route('/payment/<int:submission_id>')
    @login_required
    def payment_page(submission_id):
        """显示支付页面"""
        submission = Submission.query.get_or_404(submission_id)
        
        # 验证权限
        if not current_user.get_id().startswith('admin_') and submission.user_id != current_user.id:
            flash('无权访问', 'danger')
            return redirect(url_for('dashboard'))
        
        # 获取待支付的订单
        pending_orders = PaymentOrder.query.filter_by(
            submission_id=submission_id,
            status='pending'
        ).all()
        
        if not pending_orders:
            flash('没有待支付的订单', 'info')
            return redirect(url_for('view_submission', submission_id=submission_id))
        
        return render_template('user/payment.html', submission=submission, payment_orders=pending_orders)
    
    @app.route('/payment/process/<int:order_id>/<payment_method>')
    @login_required
    def process_payment(order_id, payment_method):
        """处理支付请求 - 使用真实支付接口"""
        payment_order = PaymentOrder.query.get_or_404(order_id)
        
        # 验证权限
        if not current_user.get_id().startswith('admin_') and payment_order.submission.user_id != current_user.id:
            flash('无权访问', 'danger')
            return redirect(url_for('dashboard'))
        
        if payment_order.status != 'pending':
            flash('订单状态异常', 'danger')
            return redirect(url_for('payment_page', submission_id=payment_order.submission_id))
        
        # 获取支付处理器
        processor = get_payment_processor()
        
        try:
            # 构建订单描述
            form_title = payment_order.submission.form.title if payment_order.submission.form else '未知表单'
            description = f"{form_title} - {payment_order.field_name}"
            
            # 根据支付方式调用相应的API
            if payment_method == 'wechat' and payment_order.payment_type == 'wechat_pay':
                result = processor.create_wechat_payment(
                    order_no=payment_order.order_no,
                    amount=float(payment_order.amount),
                    description=description
                )
            elif payment_method == 'alipay' and payment_order.payment_type == 'alipay':
                result = processor.create_alipay_payment(
                    order_no=payment_order.order_no,
                    amount=float(payment_order.amount),
                    description=description
                )
            else:
                flash('支付方式不匹配', 'danger')
                return redirect(url_for('payment_page', submission_id=payment_order.submission_id))
            
            if result.success:
                # 更新订单状态为处理中
                payment_order.status = 'processing'
                if result.trade_no:
                    payment_order.trade_no = result.trade_no
                
                # 保存支付相关数据
                payment_data = {
                    'payment_method': payment_method,
                    'api_result': result.data,
                    'payment_url': result.payment_url,
                    'created_time': datetime.utcnow().isoformat()
                }
                payment_order.set_payment_data(payment_data)
                
                db.session.commit()
                
                app.logger.info(f"💳 支付订单创建成功: 订单{payment_order.order_no}, 支付方式{payment_method}")
                
                # 根据支付方式进行跳转
                if payment_method == 'wechat':
                    # 微信支付：展示二维码或支付信息
                    return render_template('user/payment_process.html', 
                                         payment_order=payment_order, 
                                         payment_result=result,
                                         payment_method='wechat')
                elif payment_method == 'alipay':
                    # 支付宝：直接跳转到支付页面
                    if result.payment_url:
                        return redirect(result.payment_url)
                    else:
                        flash('获取支付链接失败', 'danger')
                        return redirect(url_for('payment_page', submission_id=payment_order.submission_id))
            else:
                # 支付订单创建失败
                app.logger.error(f"❌ 支付订单创建失败: {result.message}")
                flash(f'支付处理失败: {result.message}', 'danger')
                return redirect(url_for('payment_page', submission_id=payment_order.submission_id))
                
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"❌ 支付处理异常: {str(e)}")
            flash('支付处理异常，请重试', 'danger')
            return redirect(url_for('payment_page', submission_id=payment_order.submission_id))
    
    @app.route('/payment/success/<int:order_id>')
    @login_required
    def payment_success(order_id):
        """支付成功页面"""
        payment_order = PaymentOrder.query.get_or_404(order_id)
        
        # 验证权限
        if not current_user.get_id().startswith('admin_') and payment_order.submission.user_id != current_user.id:
            flash('无权访问', 'danger')
            return redirect(url_for('dashboard'))
        
        return render_template('user/payment_success.html', payment_order=payment_order)
    
    @app.route('/payment/wechat/notify', methods=['POST'])
    def wechat_payment_notify():
        """微信支付回调处理"""
        try:
            # 获取回调数据
            callback_data = request.get_data(as_text=True)
            
            # 解析XML数据
            import xml.etree.ElementTree as ET
            root = ET.fromstring(callback_data)
            data = {child.tag: child.text for child in root}
            
            app.logger.info(f"📱 收到微信支付回调: {data.get('out_trade_no')}")
            
            # 验证回调签名
            processor = get_payment_processor()
            is_valid, result = processor.verify_wechat_callback(data)
            
            if is_valid and data.get('return_code') == 'SUCCESS' and data.get('result_code') == 'SUCCESS':
                # 查找对应的支付订单
                order_no = data.get('out_trade_no')
                payment_order = PaymentOrder.query.filter_by(order_no=order_no).first()
                
                if payment_order and payment_order.status in ['pending', 'processing']:
                    # 更新订单状态
                    payment_order.status = 'paid'
                    payment_order.trade_no = data.get('transaction_id')
                    payment_order.paid_at = datetime.utcnow()
                    
                    # 保存回调数据
                    callback_info = {
                        'callback_data': data,
                        'callback_time': datetime.utcnow().isoformat(),
                        'payment_method': 'wechat'
                    }
                    existing_data = payment_order.get_payment_data() or {}
                    existing_data.update(callback_info)
                    payment_order.set_payment_data(existing_data)
                    
                    db.session.commit()
                    
                    app.logger.info(f"✅ 微信支付成功: 订单{order_no}, 交易号{data.get('transaction_id')}")
                    
                    # 返回成功响应
                    return '<xml><return_code><![CDATA[SUCCESS]]></return_code><return_msg><![CDATA[OK]]></return_msg></xml>'
                else:
                    app.logger.warning(f"⚠️ 微信支付回调: 未找到订单或状态异常 {order_no}")
            else:
                app.logger.warning(f"⚠️ 微信支付回调验证失败: {result}")
            
            # 返回失败响应
            return '<xml><return_code><![CDATA[FAIL]]></return_code><return_msg><![CDATA[FAIL]]></return_msg></xml>'
            
        except Exception as e:
            app.logger.error(f"❌ 微信支付回调处理异常: {str(e)}")
            return '<xml><return_code><![CDATA[FAIL]]></return_code><return_msg><![CDATA[ERROR]]></return_msg></xml>'
    
    @app.route('/payment/alipay/notify', methods=['POST'])
    def alipay_payment_notify():
        """支付宝支付回调处理"""
        try:
            # 获取回调数据
            callback_data = request.form.to_dict()
            
            app.logger.info(f"🐜 收到支付宝支付回调: {callback_data.get('out_trade_no')}")
            
            # 验证回调签名
            processor = get_payment_processor()
            is_valid, result = processor.verify_alipay_callback(callback_data.copy())
            
            if is_valid and callback_data.get('trade_status') == 'TRADE_SUCCESS':
                # 查找对应的支付订单
                order_no = callback_data.get('out_trade_no')
                payment_order = PaymentOrder.query.filter_by(order_no=order_no).first()
                
                if payment_order and payment_order.status in ['pending', 'processing']:
                    # 更新订单状态
                    payment_order.status = 'paid'
                    payment_order.trade_no = callback_data.get('trade_no')
                    payment_order.paid_at = datetime.utcnow()
                    
                    # 保存回调数据
                    callback_info = {
                        'callback_data': callback_data,
                        'callback_time': datetime.utcnow().isoformat(),
                        'payment_method': 'alipay'
                    }
                    existing_data = payment_order.get_payment_data() or {}
                    existing_data.update(callback_info)
                    payment_order.set_payment_data(existing_data)
                    
                    db.session.commit()
                    
                    app.logger.info(f"✅ 支付宝支付成功: 订单{order_no}, 交易号{callback_data.get('trade_no')}")
                    
                    # 返回成功响应
                    return 'success'
                else:
                    app.logger.warning(f"⚠️ 支付宝支付回调: 未找到订单或状态异常 {order_no}")
            else:
                app.logger.warning(f"⚠️ 支付宝支付回调验证失败: {result}")
            
            # 返回失败响应
            return 'fail'
            
        except Exception as e:
            app.logger.error(f"❌ 支付宝支付回调处理异常: {str(e)}")
            return 'fail'
    
    @app.route('/payment/return')
    def payment_return():
        """支付返回页面处理（支付宝同步返回）"""
        try:
            # 获取返回参数
            return_data = request.args.to_dict()
            
            if 'out_trade_no' in return_data:
                order_no = return_data.get('out_trade_no')
                payment_order = PaymentOrder.query.filter_by(order_no=order_no).first()
                
                if payment_order:
                    app.logger.info(f"🔄 支付返回: 订单{order_no}")
                    return redirect(url_for('payment_success', order_id=payment_order.id))
                else:
                    flash('订单不存在', 'danger')
                    return redirect(url_for('dashboard'))
            else:
                flash('参数错误', 'danger')
                return redirect(url_for('dashboard'))
                
        except Exception as e:
            app.logger.error(f"❌ 支付返回处理异常: {str(e)}")
            flash('支付状态异常', 'danger')
            return redirect(url_for('dashboard'))
    
    @app.route('/payment/query/<int:order_id>')
    @login_required
    def query_payment_status(order_id):
        """查询支付状态API"""
        payment_order = PaymentOrder.query.get_or_404(order_id)
        
        # 验证权限
        if not current_user.get_id().startswith('admin_') and payment_order.submission.user_id != current_user.id:
            return jsonify({'error': '无权访问'}), 403
        
        try:
            processor = get_payment_processor()
            
            # 根据支付类型查询状态
            if payment_order.payment_type == 'wechat_pay':
                result = processor.query_wechat_payment(payment_order.order_no)
            elif payment_order.payment_type == 'alipay':
                result = processor.query_alipay_payment(payment_order.order_no)
            else:
                return jsonify({'error': '不支持的支付类型'}), 400
            
            if result.success:
                # 更新本地订单状态（如果需要）
                api_data = result.data
                if api_data:
                    # 根据返回的状态更新本地订单
                    if payment_order.payment_type == 'wechat_pay':
                        trade_state = api_data.get('trade_state')
                        if trade_state == 'SUCCESS' and payment_order.status != 'paid':
                            payment_order.status = 'paid'
                            payment_order.trade_no = api_data.get('transaction_id')
                            payment_order.paid_at = datetime.utcnow()
                            db.session.commit()
                    elif payment_order.payment_type == 'alipay':
                        trade_status = api_data.get('trade_status')
                        if trade_status == 'TRADE_SUCCESS' and payment_order.status != 'paid':
                            payment_order.status = 'paid'
                            payment_order.trade_no = api_data.get('trade_no')
                            payment_order.paid_at = datetime.utcnow()
                            db.session.commit()
                
                return jsonify({
                    'success': True,
                    'status': payment_order.status,
                    'trade_no': payment_order.trade_no,
                    'api_data': api_data
                })
            else:
                return jsonify({
                    'success': False,
                    'message': result.message,
                    'error_code': result.error_code
                })
                
        except Exception as e:
            app.logger.error(f"❌ 查询支付状态异常: {str(e)}")
            return jsonify({'error': f'查询异常: {str(e)}'}), 500
    
    # 管理员路由
    @app.route('/admin/login', methods=['GET', 'POST'])
    def admin_login():
        if current_user.is_authenticated and current_user.get_id().startswith('admin_'):
            return redirect(url_for('admin_dashboard'))
        
        form = AdminLoginForm()
        if form.validate_on_submit():
            admin = Admin.query.filter_by(email=form.email.data).first()
            if admin and admin.check_password(form.password.data):
                login_user(admin)
                return redirect(url_for('admin_dashboard'))
            flash('邮箱或密码错误', 'danger')
        
        return render_template('admin/login.html', form=form)
    
    @app.route('/admin/dashboard')
    @login_required
    def admin_dashboard():
        if not current_user.get_id().startswith('admin_'):
            flash('无权访问管理后台', 'danger')
            return redirect(url_for('index'))
        
        forms = Form.query.order_by(Form.created_at.desc()).all()
        users = User.query.order_by(User.created_at.desc()).limit(10).all()
        
        # 统计数据
        stats = {
            'total_forms': Form.query.count(),
            'total_users': User.query.count(),
            'total_submissions': Submission.query.count(),
            'active_forms': Form.query.filter_by(is_active=True).count(),
            'total_payments': PaymentOrder.query.count(),
            'paid_orders': PaymentOrder.query.filter_by(status='paid').count(),
            'pending_payments': PaymentOrder.query.filter_by(status='pending').count(),
            'total_revenue': db.session.query(db.func.sum(PaymentOrder.amount)).filter_by(status='paid').scalar() or 0
        }
        
        return render_template('admin/dashboard.html', forms=forms, users=users, stats=stats)
    
    @app.route('/admin/forms')
    @login_required
    def admin_forms():
        if not current_user.get_id().startswith('admin_'):
            flash('无权访问', 'danger')
            return redirect(url_for('index'))
        
        forms = Form.query.order_by(Form.created_at.desc()).all()
        return render_template('admin/forms.html', forms=forms)
    
    @app.route('/admin/forms/create', methods=['GET', 'POST'])
    @login_required
    def admin_create_form():
        if not current_user.get_id().startswith('admin_'):
            flash('无权访问', 'danger')
            return redirect(url_for('index'))
        
        if request.method == 'POST':
            # 处理表单创建
            title = request.form.get('title')
            description = request.form.get('description')
            
            if not title:
                flash('请填写表单标题', 'danger')
                return render_template('admin/create_form.html')
            
            # 创建表单
            admin_id = int(current_user.get_id().replace('admin_', ''))
            form_obj = Form(
                title=title,
                description=description,
                created_by=admin_id
            )
            db.session.add(form_obj)
            db.session.flush()
            
            # 添加字段
            field_count = int(request.form.get('field_count', 0))
            for i in range(field_count):
                field_name = request.form.get(f'field_{i}_name')
                field_label = request.form.get(f'field_{i}_label')
                field_type = request.form.get(f'field_{i}_type')
                is_required = request.form.get(f'field_{i}_required') == 'on'
                placeholder = request.form.get(f'field_{i}_placeholder', '')
                options = request.form.get(f'field_{i}_options', '')
                payment_account_id = request.form.get(f'field_{i}_payment_account') or None
                
                if field_name and field_label and field_type:
                    form_field = FormField(
                        form_id=form_obj.id,
                        field_name=field_name,
                        field_label=field_label,
                        field_type=field_type,
                        is_required=is_required,
                        placeholder=placeholder,
                        order_index=i,
                        payment_account_id=int(payment_account_id) if payment_account_id else None
                    )
                    
                    if options and field_type in ['select', 'radio', 'checkbox']:
                        options_list = [opt.strip() for opt in options.split('\n') if opt.strip()]
                        form_field.set_options(options_list)
                    
                    db.session.add(form_field)
            
            db.session.commit()
            flash('表单创建成功！', 'success')
            return redirect(url_for('admin_forms'))
        
        return render_template('admin/create_form.html')
    
    @app.route('/admin/forms/<int:form_id>/edit', methods=['GET', 'POST'])
    @login_required
    def admin_edit_form(form_id):
        if not current_user.get_id().startswith('admin_'):
            flash('无权访问', 'danger')
            return redirect(url_for('index'))
        
        form_obj = Form.query.get_or_404(form_id)
        
        if request.method == 'POST':
            form_obj.title = request.form.get('title', form_obj.title)
            form_obj.description = request.form.get('description', form_obj.description)
            form_obj.is_active = request.form.get('is_active') == 'on'
            form_obj.updated_at = datetime.utcnow()
            
            db.session.commit()
            flash('表单更新成功！', 'success')
            return redirect(url_for('admin_forms'))
        
        return render_template('admin/edit_form.html', form=form_obj)
    
    @app.route('/admin/forms/<int:form_id>/submissions')
    @login_required
    def admin_form_submissions(form_id):
        if not current_user.get_id().startswith('admin_'):
            flash('无权访问', 'danger')
            return redirect(url_for('index'))
        
        form_obj = Form.query.get_or_404(form_id)
        submissions = Submission.query.filter_by(form_id=form_id).order_by(Submission.submitted_at.desc()).all()
        
        return render_template('admin/form_submissions.html', form=form_obj, submissions=submissions)
    
    @app.route('/admin/users')
    @login_required
    def admin_users():
        if not current_user.get_id().startswith('admin_'):
            flash('无权访问', 'danger')
            return redirect(url_for('index'))
        
        users = User.query.order_by(User.created_at.desc()).all()
        return render_template('admin/users.html', users=users)
    
    @app.route('/admin/payments')
    @login_required
    def admin_payments():
        """支付订单管理"""
        if not current_user.get_id().startswith('admin_'):
            flash('无权访问', 'danger')
            return redirect(url_for('index'))
        
        # 获取筛选参数
        status_filter = request.args.get('status', '')
        payment_type_filter = request.args.get('payment_type', '')
        
        # 构建查询
        query = PaymentOrder.query
        
        if status_filter:
            query = query.filter(PaymentOrder.status == status_filter)
        if payment_type_filter:
            query = query.filter(PaymentOrder.payment_type == payment_type_filter)
        
        payments = query.order_by(PaymentOrder.created_at.desc()).all()
        
        # 统计数据
        payment_stats = {
            'total_orders': PaymentOrder.query.count(),
            'paid_orders': PaymentOrder.query.filter_by(status='paid').count(),
            'pending_orders': PaymentOrder.query.filter_by(status='pending').count(),
            'failed_orders': PaymentOrder.query.filter_by(status='failed').count(),
            'total_revenue': db.session.query(db.func.sum(PaymentOrder.amount)).filter_by(status='paid').scalar() or 0,
            'wechat_revenue': db.session.query(db.func.sum(PaymentOrder.amount)).filter_by(status='paid', payment_type='wechat_pay').scalar() or 0,
            'alipay_revenue': db.session.query(db.func.sum(PaymentOrder.amount)).filter_by(status='paid', payment_type='alipay').scalar() or 0
        }
        
        return render_template('admin/payments.html', payments=payments, payment_stats=payment_stats)
    
    @app.route('/admin/payments/<int:order_id>/update-status', methods=['POST'])
    @login_required
    def admin_update_payment_status(order_id):
        """更新支付订单状态"""
        if not current_user.get_id().startswith('admin_'):
            return jsonify({'error': '无权访问'}), 403
        
        try:
            payment_order = PaymentOrder.query.get_or_404(order_id)
            
            # 获取新状态
            data = request.get_json() if request.is_json else request.form
            new_status = data.get('status')
            
            if new_status not in ['pending', 'paid', 'failed', 'cancelled']:
                return jsonify({'error': '无效的状态值'}), 400
            
            # 更新状态
            old_status = payment_order.status
            payment_order.status = new_status
            
            # 如果是标记为已支付，记录支付时间
            if new_status == 'paid' and old_status != 'paid':
                payment_order.paid_at = datetime.utcnow()
                if not payment_order.trade_no:
                    payment_order.trade_no = f"ADMIN{int(time.time())}{random.randint(100, 999)}"
            
            db.session.commit()
            
            # 返回成功响应
            status_text = {
                'pending': '待支付',
                'paid': '已支付', 
                'failed': '支付失败',
                'cancelled': '已取消'
            }
            
            return jsonify({
                'success': True, 
                'message': f'状态已从"{status_text.get(old_status, old_status)}"更新为"{status_text.get(new_status)}"',
                'new_status': new_status,
                'status_text': status_text.get(new_status)
            })
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"更新支付状态失败: {str(e)}")
            return jsonify({'error': '状态更新失败'}), 500
    
    @app.route('/admin/payment-accounts')
    @login_required
    def admin_payment_accounts():
        """收款账户管理"""
        if not current_user.get_id().startswith('admin_'):
            flash('无权访问', 'danger')
            return redirect(url_for('index'))
        
        accounts = PaymentAccount.query.order_by(PaymentAccount.created_at.desc()).all()
        return render_template('admin/payment_accounts.html', accounts=accounts)
    
    @app.route('/admin/payment-accounts/create', methods=['GET', 'POST'])
    @login_required
    def admin_create_payment_account():
        """创建收款账户"""
        if not current_user.get_id().startswith('admin_'):
            flash('无权访问', 'danger')
            return redirect(url_for('index'))
        
        if request.method == 'POST':
            try:
                account_name = request.form.get('account_name')
                account_type = request.form.get('account_type')
                account_number = request.form.get('account_number')
                account_holder = request.form.get('account_holder')
                bank_name = request.form.get('bank_name', '')
                bank_branch = request.form.get('bank_branch', '')
                notes = request.form.get('notes', '')
                
                if not all([account_name, account_type, account_number, account_holder]):
                    flash('请填写必要信息', 'danger')
                    return render_template('admin/create_payment_account.html')
                
                # 创建收款账户
                admin_id = int(current_user.get_id().replace('admin_', ''))
                payment_account = PaymentAccount(
                    account_name=account_name,
                    account_type=account_type,
                    account_number=account_number,
                    account_holder=account_holder,
                    bank_name=bank_name if account_type == 'bank_card' else None,
                    bank_branch=bank_branch if account_type == 'bank_card' else None,
                    notes=notes,
                    created_by=admin_id
                )
                
                db.session.add(payment_account)
                db.session.commit()
                
                flash('收款账户创建成功！', 'success')
                return redirect(url_for('admin_payment_accounts'))
                
            except Exception as e:
                db.session.rollback()
                app.logger.error(f"创建收款账户失败: {str(e)}")
                flash('创建失败，请重试', 'danger')
        
        return render_template('admin/create_payment_account.html')
    
    @app.route('/admin/payment-accounts/<int:account_id>/edit', methods=['GET', 'POST'])
    @login_required
    def admin_edit_payment_account(account_id):
        """编辑收款账户"""
        if not current_user.get_id().startswith('admin_'):
            flash('无权访问', 'danger')
            return redirect(url_for('index'))
        
        account = PaymentAccount.query.get_or_404(account_id)
        
        if request.method == 'POST':
            try:
                account.account_name = request.form.get('account_name', account.account_name)
                account.account_number = request.form.get('account_number', account.account_number)
                account.account_holder = request.form.get('account_holder', account.account_holder)
                account.bank_name = request.form.get('bank_name', account.bank_name)
                account.bank_branch = request.form.get('bank_branch', account.bank_branch)
                account.notes = request.form.get('notes', account.notes)
                account.is_active = request.form.get('is_active') == 'on'
                
                db.session.commit()
                flash('账户信息更新成功！', 'success')
                return redirect(url_for('admin_payment_accounts'))
                
            except Exception as e:
                db.session.rollback()
                app.logger.error(f"更新收款账户失败: {str(e)}")
                flash('更新失败，请重试', 'danger')
        
        return render_template('admin/edit_payment_account.html', account=account)
    
    @app.route('/admin/payment-accounts/<int:account_id>/toggle-status', methods=['POST'])
    @login_required
    def admin_toggle_payment_account_status(account_id):
        """切换收款账户状态"""
        if not current_user.get_id().startswith('admin_'):
            return jsonify({'error': '无权访问'}), 403
        
        try:
            account = PaymentAccount.query.get_or_404(account_id)
            account.is_active = not account.is_active
            db.session.commit()
            
            return jsonify({
                'success': True,
                'is_active': account.is_active,
                'message': f'账户 {account.account_name} 已{"启用" if account.is_active else "禁用"}'
            })
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"账户状态切换失败: {str(e)}")
            return jsonify({'error': '状态更新失败'}), 500
    
    @app.route('/admin/database/clear', methods=['GET', 'POST'])
    @login_required
    def admin_clear_database():
        """清空数据库功能"""
        if not current_user.get_id().startswith('admin_'):
            flash('无权访问', 'danger')
            return redirect(url_for('index'))
        
        if request.method == 'POST':
            # 验证确认码
            confirm_code = request.form.get('confirm_code')
            if confirm_code != 'CLEAR_ALL_DATA':
                flash('确认码错误，操作已取消', 'danger')
                return render_template('admin/clear_database.html')
            
            try:
                app.logger.warning(f"🚨 管理员 {current_user.email} 正在清空数据库")
                
                # 删除数据但保留当前管理员账户
                admin_id = int(current_user.get_id().replace('admin_', ''))
                
                # 按顺序清空表数据（考虑外键约束）
                from sqlalchemy import text
                db.session.execute(text('DELETE FROM submission_data'))
                db.session.execute(text('DELETE FROM upload_file'))
                db.session.execute(text('DELETE FROM payment_order'))
                db.session.execute(text('DELETE FROM submission'))
                db.session.execute(text('DELETE FROM form_field'))
                db.session.execute(text('DELETE FROM form'))
                db.session.execute(text('DELETE FROM payment_account'))
                db.session.execute(text('DELETE FROM user'))
                # 只保留当前登录的管理员
                db.session.execute(text('DELETE FROM admin WHERE id != :admin_id'), {'admin_id': admin_id})
                
                db.session.commit()
                
                # 清理上传文件
                try:
                    upload_path = os.path.join(app.instance_path, app.config['UPLOAD_FOLDER'])
                    if os.path.exists(upload_path):
                        for filename in os.listdir(upload_path):
                            file_path = os.path.join(upload_path, filename)
                            if os.path.isfile(file_path):
                                os.remove(file_path)
                        app.logger.info("📁 上传文件已清理")
                except Exception as e:
                    app.logger.warning(f"清理上传文件时出错: {str(e)}")
                
                app.logger.warning("🧹 数据库已清空（保留当前管理员）")
                flash('数据库已成功清空！当前管理员账户已保留。', 'success')
                return redirect(url_for('admin_dashboard'))
                
            except Exception as e:
                db.session.rollback()
                app.logger.error(f"❌ 清空数据库失败: {str(e)}")
                flash('清空数据库失败，请重试', 'danger')
                return render_template('admin/clear_database.html')
        
        return render_template('admin/clear_database.html')
    
    @app.route('/admin/system/management')
    @login_required
    def admin_system_management():
        """系统管理页面"""
        if not current_user.get_id().startswith('admin_'):
            flash('无权访问', 'danger')
            return redirect(url_for('index'))
        
        # 统计数据
        system_stats = {
            'total_files': 0,
            'database_size': '0 MB',
            'last_backup': None,
            'uptime': '1 天 3 小时'  # 模拟数据
        }
        
        # 计算上传文件数量
        try:
            upload_path = os.path.join(app.instance_path, app.config['UPLOAD_FOLDER'])
            if os.path.exists(upload_path):
                system_stats['total_files'] = len([f for f in os.listdir(upload_path) if os.path.isfile(os.path.join(upload_path, f))])
        except:
            pass
        
        # 计算数据库文件大小
        try:
            db_file = os.path.join(app.instance_path, 'form_system.db')
            if os.path.exists(db_file):
                size = os.path.getsize(db_file)
                system_stats['database_size'] = f"{size / 1024 / 1024:.2f} MB"
        except:
            pass
        
        return render_template('admin/system_management.html', system_stats=system_stats)
    
    @app.route('/admin/maintenance/database/emergency-clear')
    @login_required
    def admin_emergency_clear():
        """紧急数据库清空直接访问路径"""
        if not current_user.get_id().startswith('admin_'):
            flash('无权访问', 'danger')
            return redirect(url_for('index'))
        
        # 记录访问日志
        app.logger.warning(f"🚨 管理员 {current_user.email} 通过紧急路径访问清空数据库功能")
        
        return render_template('admin/clear_database.html')
    
    @app.route('/admin/api/payment-accounts')
    @login_required
    def api_get_payment_accounts():
        """获取收款账户列表API"""
        if not current_user.get_id().startswith('admin_'):
            return jsonify({'error': '无权访问'}), 403
        
        try:
            accounts = PaymentAccount.query.filter_by(is_active=True).all()
            accounts_data = []
            
            for account in accounts:
                accounts_data.append({
                    'id': account.id,
                    'account_name': account.account_name,
                    'account_type': account.account_type,
                    'account_holder': account.account_holder,
                    'get_account_display': account.get_account_display(),
                    'is_active': account.is_active
                })
            
            return jsonify({
                'success': True,
                'accounts': accounts_data
            })
            
        except Exception as e:
            app.logger.error(f"获取收款账户列表失败: {str(e)}")
            return jsonify({'error': '获取失败'}), 500
    
    @app.route('/admin/test-export')
    @login_required
    def test_export():
        """测试导出功能"""
        if not current_user.get_id().startswith('admin_'):
            return jsonify({'error': '无权访问'}), 403
        
        try:
            app.logger.info("开始测试导出")
            
            # 创建一个简单的Excel文件
            wb = Workbook()
            ws = wb.active
            ws.title = "测试数据"
            
            # 添加测试数据
            ws.cell(row=1, column=1, value="测试列1")
            ws.cell(row=1, column=2, value="测试列2")
            ws.cell(row=2, column=1, value="数据1")
            ws.cell(row=2, column=2, value="数据2")
            
            # 保存到内存
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 创建响应
            response = make_response(output.read())
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f'test_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response.headers['Content-Disposition'] = encode_filename_for_http(filename)
            
            app.logger.info(f"测试导出完成: {filename}")
            return response
            
        except Exception as e:
            app.logger.error(f"测试导出失败: {str(e)}", exc_info=True)
            return jsonify({'error': f'测试导出失败: {str(e)}'}), 500
    
    @app.route('/admin/export/users')
    @login_required
    def admin_export_users():
        """导出用户数据，支持多种格式和选项"""
        if not current_user.get_id().startswith('admin_'):
            return jsonify({'error': '无权访问'}), 403
        
        try:
            app.logger.info(f"开始导出用户数据，参数: {request.args}")
            
            # 获取查询参数
            format_type = request.args.get('format', 'excel')  # excel 或 csv
            file_name = request.args.get('fileName', 'users_data')
            
            # 获取导出内容选项
            include_basic = request.args.get('includeBasicInfo', 'true') == 'true'
            include_contact = request.args.get('includeContactInfo', 'true') == 'true'
            include_submissions = request.args.get('includeSubmissions', 'true') == 'true'
            include_activity = request.args.get('includeActivity', 'true') == 'true'
            
            # 获取筛选选项
            status_filter = request.args.get('statusFilter', '')
            type_filter = request.args.get('typeFilter', '')
            
            app.logger.info(f"导出选项: format={format_type}, file_name={file_name}, include_basic={include_basic}")
            
            # 构建查询
            query = User.query
            
            # 应用状态筛选
            if status_filter == 'active':
                query = query.filter(User.is_active == True)
            elif status_filter == 'inactive':
                query = query.filter(User.is_active == False)
            
            # 应用类型筛选
            if type_filter == 'email':
                query = query.filter(User.email.isnot(None))
            elif type_filter == 'phone':
                query = query.filter(User.phone.isnot(None))
            
            users = query.order_by(User.created_at.desc()).all()
            app.logger.info(f"查询到 {len(users)} 个用户")
            
            if format_type == 'csv':
                # CSV格式导出
                output = io.StringIO()
                writer = csv.writer(output)
                
                # 构建表头
                headers = []
                if include_basic:
                    headers.extend(['用户ID', '用户名', '注册时间', '账户状态'])
                if include_contact:
                    headers.extend(['邮箱', '手机号'])
                if include_submissions:
                    headers.extend(['提交数量'])
                if include_activity:
                    headers.extend(['最后活动时间', '最后提交表单'])
                
                writer.writerow(headers)
                
                # 写入数据
                for user in users:
                    row = []
                    if include_basic:
                        row.extend([
                            user.id,
                            user.name,
                            user.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                            '活跃' if user.is_active else '禁用'
                        ])
                    if include_contact:
                        row.extend([
                            user.email or '未设置',
                            user.phone or '未设置'
                        ])
                    if include_submissions:
                        row.append(len(user.submissions))
                    if include_activity:
                        last_submission = None
                        if user.submissions:
                            last_submission = max(user.submissions, key=lambda s: s.submitted_at)
                        row.extend([
                            last_submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if last_submission else '无记录',
                            (last_submission.form.title if last_submission.form else '表单已删除') if last_submission else '无记录'
                        ])
                    writer.writerow(row)
                
                # 创建响应
                csv_content = output.getvalue()
                # 添加BOM以支持中文编码
                csv_content = '\ufeff' + csv_content
                response = make_response(csv_content.encode('utf-8'))
                response.headers['Content-Type'] = 'text/csv; charset=utf-8'
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f'{file_name}_{timestamp}.csv'
                response.headers['Content-Disposition'] = encode_filename_for_http(filename)
                
                app.logger.info(f"CSV导出完成，文件名: {filename}")
                return response
            
            else:
                # Excel格式导出
                app.logger.info("开始生成Excel文件")
                wb = Workbook()
                ws = wb.active
                ws.title = "用户数据"
                
                # 设置表头样式
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # 构建表头
                headers = []
                if include_basic:
                    headers.extend(['用户ID', '用户名', '注册时间', '账户状态'])
                if include_contact:
                    headers.extend(['邮箱', '手机号'])
                if include_submissions:
                    headers.extend(['提交数量'])
                if include_activity:
                    headers.extend(['最后活动时间', '最后提交表单'])
                
                # 写入表头
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 写入用户数据
                for row, user in enumerate(users, 2):
                    col = 1
                    if include_basic:
                        ws.cell(row=row, column=col, value=user.id)
                        ws.cell(row=row, column=col+1, value=user.name)
                        ws.cell(row=row, column=col+2, value=user.created_at.strftime('%Y-%m-%d %H:%M:%S'))
                        ws.cell(row=row, column=col+3, value='活跃' if user.is_active else '禁用')
                        col += 4
                    
                    if include_contact:
                        ws.cell(row=row, column=col, value=user.email or '未设置')
                        ws.cell(row=row, column=col+1, value=user.phone or '未设置')
                        col += 2
                    
                    if include_submissions:
                        ws.cell(row=row, column=col, value=len(user.submissions))
                        col += 1
                    
                    if include_activity:
                        last_submission = None
                        if user.submissions:
                            last_submission = max(user.submissions, key=lambda s: s.submitted_at)
                        ws.cell(row=row, column=col, value=last_submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if last_submission else '无记录')
                        ws.cell(row=row, column=col+1, value=(last_submission.form.title if last_submission.form else '表单已删除') if last_submission else '无记录')
                        col += 2
                
                # 调整列宽
                for col in range(1, len(headers) + 1):
                    max_length = 0
                    column = ws.cell(row=1, column=col).column_letter
                    for row_cells in ws[column]:
                        try:
                            if len(str(row_cells.value)) > max_length:
                                max_length = len(str(row_cells.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # 最大宽度50
                    ws.column_dimensions[column].width = max(adjusted_width, 10)  # 最小宽度10
                
                app.logger.info("Excel文件生成完成，开始保存到内存")
                
                # 保存到内存
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                # 创建响应
                response = make_response(output.read())
                response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f'{file_name}_{timestamp}.xlsx'
                response.headers['Content-Disposition'] = encode_filename_for_http(filename)
                
                app.logger.info(f"Excel导出完成，文件名: {filename}")
                return response
            
        except Exception as e:
            app.logger.error(f"用户数据导出失败: {str(e)}", exc_info=True)
            return jsonify({'error': f'导出失败: {str(e)}'}), 500
    
    @app.route('/admin/test-form-export/<int:form_id>')
    @login_required
    def test_form_export(form_id):
        """测试表单导出路由"""
        if not current_user.get_id().startswith('admin_'):
            return jsonify({'error': '无权访问'}), 403
        
        app.logger.info(f"测试表单导出路由 - 表单ID: {form_id}")
        return jsonify({
            'success': True, 
            'message': f'路由正常，表单ID: {form_id}',
            'form_id': form_id,
            'args': dict(request.args)
        })
    
    @app.route('/admin/debug-export-forms/<int:form_id>')
    @login_required
    def debug_export_form(form_id):
        """调试表单导出功能"""
        if not current_user.get_id().startswith('admin_'):
            return jsonify({'error': '无权访问'}), 403
        
        try:
            app.logger.info(f"调试表单导出 - 表单ID: {form_id}, 参数: {dict(request.args)}")
            
            # 检查表单是否存在
            form_obj = Form.query.get(form_id)
            if not form_obj:
                return jsonify({
                    'success': False,
                    'error': f'表单ID {form_id} 不存在',
                    'form_id': form_id
                }), 404
            
            # 检查提交记录数量
            submissions_count = Submission.query.filter_by(form_id=form_id).count()
            
            # 检查表单字段
            fields_count = FormField.query.filter_by(form_id=form_id).count()
            
            return jsonify({
                'success': True,
                'message': '调试信息获取成功',
                'form_id': form_id,
                'form_title': form_obj.title,
                'submissions_count': submissions_count,
                'fields_count': fields_count,
                'request_args': dict(request.args),
                'url_path': request.path,
                'full_url': request.full_path
            })
            
        except Exception as e:
            app.logger.error(f"调试表单导出失败: {str(e)}", exc_info=True)
            return jsonify({
                'success': False,
                'error': str(e),
                'form_id': form_id
            }), 500
    
    @app.route('/admin/export/forms/<int:form_id>')
    @login_required
    def admin_export_form_submissions(form_id):
        """导出表单提交数据，支持多种格式和选项"""
        try:
            app.logger.info(f"🚀 导出请求开始 - 表单ID: {form_id}, 参数: {dict(request.args)}")
            app.logger.info(f"👤 当前用户: {current_user.get_id() if current_user and current_user.is_authenticated else 'None'}")
            app.logger.info(f"🌍 请求URL: {request.url}")
            app.logger.info(f"📡 请求方法: {request.method}")
            
            # 立即输出到控制台确保日志可见
            print(f"\n🚀 ==========导出请求开始========== ")
            print(f"👤 当前用户: {current_user.get_id() if current_user and current_user.is_authenticated else 'None'}")
            print(f"🌍 请求URL: {request.url}")
            print(f"📡 请求方法: {request.method}")
            print(f"📋 参数: {dict(request.args)}")
            print(f"🚀 ============================================ \n")
            
            if not current_user or not current_user.is_authenticated:
                app.logger.error("❌ 用户未登录")
                return jsonify({'error': '请先登录'}), 401
                
            if not current_user.get_id().startswith('admin_'):
                app.logger.warning(f"⚠️ 非管理员用户尝试访问导出功能: {current_user.get_id()}")
                return jsonify({'error': '无权访问，需要管理员权限'}), 403
            app.logger.info(f"📋 开始导出表单提交数据，表单ID: {form_id}，参数: {request.args}")
            
            # 检查表单是否存在
            form_obj = Form.query.get(form_id)
            if not form_obj:
                app.logger.error(f"❌ 表单不存在: {form_id}")
                return jsonify({'error': f'表单ID {form_id} 不存在'}), 404
            
            app.logger.info(f"✅ 表单信息: {form_obj.title}")
            
            # 获取查询参数
            format_type = request.args.get('format', 'excel')  # excel 或 csv 或 zip
            file_name = request.args.get('fileName', f'{form_obj.title}_submissions')
            
            # 获取导出内容选项
            include_submitter = request.args.get('includeSubmitterInfo', 'true') == 'true'
            include_time = request.args.get('includeSubmissionTime', 'true') == 'true'
            include_form_data = request.args.get('includeFormData', 'true') == 'true'
            include_attachments = request.args.get('includeAttachments', 'true') == 'true'
            
            # 获取状态筛选
            status_filter = request.args.get('statusFilter', '')
            
            app.logger.info(f"📊 导出选项: format={format_type}, file_name={file_name}, include_submitter={include_submitter}")
            
            # 构建查询
            query = Submission.query.filter_by(form_id=form_id)
            
            # 应用状态筛选
            if status_filter:
                query = query.filter(Submission.status == status_filter)
                app.logger.info(f"🔍 应用状态筛选: {status_filter}")
            
            submissions = query.order_by(Submission.submitted_at.desc()).all()
            app.logger.info(f"📝 查询到 {len(submissions)} 个提交记录")
            
            # 获取表单字段
            fields = FormField.query.filter_by(form_id=form_id).order_by(FormField.order_index).all()
            app.logger.info(f"📋 表单字段数量: {len(fields)}")
            
            if format_type == 'csv':
                app.logger.info("📄 开始生成CSV文件")
                return export_csv(submissions, fields, file_name, include_submitter, include_time, include_form_data, include_attachments)
            
            elif format_type == 'zip':
                app.logger.info("📦 开始生成ZIP打包文件")
                return export_zip(form_obj, submissions, fields, file_name, include_submitter, include_time, include_form_data, include_attachments)
            
            else:
                app.logger.info("📊 开始生成Excel文件")
                return export_excel(form_obj, submissions, fields, file_name, include_submitter, include_time, include_form_data, include_attachments)
            
        except Exception as e:
            app.logger.error(f"❌ 表单数据导出失败: {str(e)}", exc_info=True)
            return jsonify({'error': f'导出失败: {str(e)}'}), 500
    
    def export_csv(submissions, fields, file_name, include_submitter, include_time, include_form_data, include_attachments):
        """导出CSV格式"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 构建表头
        headers = ['提交ID']
        if include_submitter:
            headers.extend(['提交者', '邮箱', '手机号'])
        if include_time:
            headers.extend(['提交时间', '状态'])
        if include_form_data:
            for field in fields:
                headers.append(field.field_label)
        if include_attachments:
            headers.append('附件信息')
        
        writer.writerow(headers)
        app.logger.info(f"📋 CSV表头: {headers}")
        
        # 写入数据
        for i, submission in enumerate(submissions):
            try:
                data_dict = submission.get_data_dict()
                
                # 状态映射
                status_map = {
                    'submitted': '待审核',
                    'approved': '已通过',
                    'rejected': '已拒绝'
                }
                
                row = [submission.id]
                if include_submitter:
                    row.extend([
                        submission.user.name if submission.user else '未知用户',
                        submission.user.email or '未设置' if submission.user else '未设置',
                        submission.user.phone or '未设置' if submission.user else '未设置'
                    ])
                if include_time:
                    row.extend([
                        submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
                        status_map.get(submission.status, submission.status)
                    ])
                if include_form_data:
                    for field in fields:
                        field_value = data_dict.get(field.field_name, '')
                        # 处理复选框数据
                        if field.field_type == 'checkbox' and ',' in str(field_value):
                            field_value = field_value.replace(',', '; ')
                        row.append(field_value)
                if include_attachments:
                    # 构建附件信息（包含下载链接）
                    attachment_info = []
                    if submission.files:
                        for file in submission.files:
                            # 生成文件下载链接
                            download_url = f"{request.host_url}uploads/{file.saved_filename}"
                            file_info = f"{file.original_filename} ({download_url})"
                            attachment_info.append(file_info)
                    
                    # 将附件信息合并为一个字符串
                    attachment_text = '; '.join(attachment_info) if attachment_info else '无附件'
                    row.append(attachment_text)
                
                writer.writerow(row)
                
                if (i + 1) % 100 == 0:  # 每100条记录记录一次日志
                    app.logger.info(f"📝 已处理 {i + 1} 条记录")
                    
            except Exception as row_error:
                app.logger.error(f"❌ 处理第 {i+1} 条记录时出错: {str(row_error)}", exc_info=True)
                # 继续处理下一条记录，不中断整个导出过程
                continue
        
        # 创建响应
        csv_content = output.getvalue()
        # 添加BOM以支持中文编码
        csv_content = '\ufeff' + csv_content
        response = make_response(csv_content.encode('utf-8'))
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'{file_name}_{timestamp}.csv'
        response.headers['Content-Disposition'] = encode_filename_for_http(filename)
        
        app.logger.info(f"✅ CSV导出完成，文件名: {filename}, 大小: {len(csv_content)} bytes")
        return response
    
    def export_excel(form_obj, submissions, fields, file_name, include_submitter, include_time, include_form_data, include_attachments):
        """导出Excel格式"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"{form_obj.title}提交数据"
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 构建表头
        headers = ['提交ID']
        if include_submitter:
            headers.extend(['提交者', '邮箱', '手机号'])
        if include_time:
            headers.extend(['提交时间', '状态'])
        if include_form_data:
            for field in fields:
                headers.append(field.field_label)
        if include_attachments:
            headers.append('附件信息')
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        app.logger.info(f"📋 Excel表头: {headers}")
        
        # 写入提交数据
        for row_num, submission in enumerate(submissions, 2):
            try:
                data_dict = submission.get_data_dict()
                
                # 状态映射
                status_map = {
                    'submitted': '待审核',
                    'approved': '已通过',
                    'rejected': '已拒绝'
                }
                
                col = 1
                ws.cell(row=row_num, column=col, value=submission.id)
                col += 1
                
                if include_submitter:
                    ws.cell(row=row_num, column=col, value=submission.user.name if submission.user else '未知用户')
                    ws.cell(row=row_num, column=col+1, value=submission.user.email or '未设置' if submission.user else '未设置')
                    ws.cell(row=row_num, column=col+2, value=submission.user.phone or '未设置' if submission.user else '未设置')
                    col += 3
                
                if include_time:
                    ws.cell(row=row_num, column=col, value=submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row=row_num, column=col+1, value=status_map.get(submission.status, submission.status))
                    col += 2
                
                if include_form_data:
                    for field in fields:
                        field_value = data_dict.get(field.field_name, '')
                        # 处理复选框数据
                        if field.field_type == 'checkbox' and ',' in str(field_value):
                            field_value = field_value.replace(',', '; ')
                        ws.cell(row=row_num, column=col, value=field_value)
                        col += 1
                
                if include_attachments:
                    # 构建附件信息（包含下载链接）
                    attachment_info = []
                    if submission.files:
                        for file in submission.files:
                            # 生成文件下载链接
                            download_url = f"{request.host_url}uploads/{file.saved_filename}"
                            file_info = f"{file.original_filename} ({download_url})"
                            attachment_info.append(file_info)
                    
                    # 将附件信息合并为一个字符串
                    attachment_text = '; '.join(attachment_info) if attachment_info else '无附件'
                    ws.cell(row=row_num, column=col, value=attachment_text)
                    col += 1
                
                if row_num % 100 == 1:  # 每100条记录记录一次日志
                    app.logger.info(f"📝 已处理 {row_num - 1} 条Excel记录")
            
            except Exception as row_error:
                app.logger.error(f"❌ 处理Excel第 {row_num-1} 条记录时出错: {str(row_error)}", exc_info=True)
                # 继续处理下一条记录
                continue
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = ws.cell(row=1, column=col).column_letter
            for row_cells in ws[column]:
                try:
                    if len(str(row_cells.value)) > max_length:
                        max_length = len(str(row_cells.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 最大宽度50
            ws.column_dimensions[column].width = max(adjusted_width, 10)  # 最小宽度10
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 创建响应
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'{file_name}_{timestamp}.xlsx'
        response.headers['Content-Disposition'] = encode_filename_for_http(filename)
        
        app.logger.info(f"✅ Excel导出完成，文件名: {filename}")
        return response
    
    def export_zip(form_obj, submissions, fields, file_name, include_submitter, include_time, include_form_data, include_attachments):
        """导出ZIP打包格式（包含Excel文件和所有附件）"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 创建临时目录
        temp_dir = os.path.join(upload_dir, f'temp_export_{timestamp}')
        os.makedirs(temp_dir, exist_ok=True)
        
        try:
            # 首先生成Excel文件
            wb = Workbook()
            ws = wb.active
            ws.title = f"{form_obj.title}提交数据"
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 构建表头
            headers = ['提交ID']
            if include_submitter:
                headers.extend(['提交者', '邮箱', '手机号'])
            if include_time:
                headers.extend(['提交时间', '状态'])
            if include_form_data:
                for field in fields:
                    headers.append(field.field_label)
            if include_attachments:
                headers.append('附件信息')
            
            # 写入表头
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            app.logger.info(f"📋 Excel表头: {headers}")
            
            # 创建附件目录
            attachments_dir = os.path.join(temp_dir, '附件')
            os.makedirs(attachments_dir, exist_ok=True)
            
            # 处理数据和附件
            attachment_files = []
            for row_num, submission in enumerate(submissions, 2):
                try:
                    data_dict = submission.get_data_dict()
                    
                    # 状态映射
                    status_map = {
                        'submitted': '待审核',
                        'approved': '已通过',
                        'rejected': '已拒绝'
                    }
                    
                    col = 1
                    ws.cell(row=row_num, column=col, value=submission.id)
                    col += 1
                    
                    if include_submitter:
                        ws.cell(row=row_num, column=col, value=submission.user.name if submission.user else '未知用户')
                        ws.cell(row=row_num, column=col+1, value=submission.user.email or '未设置' if submission.user else '未设置')
                        ws.cell(row=row_num, column=col+2, value=submission.user.phone or '未设置' if submission.user else '未设置')
                        col += 3
                    
                    if include_time:
                        ws.cell(row=row_num, column=col, value=submission.submitted_at.strftime('%Y-%m-%d %H:%M:%S'))
                        ws.cell(row=row_num, column=col+1, value=status_map.get(submission.status, submission.status))
                        col += 2
                    
                    if include_form_data:
                        for field in fields:
                            field_value = data_dict.get(field.field_name, '')
                            # 处理复选框数据
                            if field.field_type == 'checkbox' and ',' in str(field_value):
                                field_value = field_value.replace(',', '; ')
                            ws.cell(row=row_num, column=col, value=field_value)
                            col += 1
                    
                    if include_attachments:
                        # 处理附件
                        attachment_info = []
                        if submission.files:
                            for file in submission.files:
                                # 复制附件到临时目录
                                source_path = os.path.join(upload_dir, file.saved_filename)
                                if os.path.exists(source_path):
                                    # 使用提交ID和原始文件名创建新文件名
                                    new_filename = f"提交{submission.id}_{file.original_filename}"
                                    dest_path = os.path.join(attachments_dir, new_filename)
                                    shutil.copy2(source_path, dest_path)
                                    attachment_files.append(new_filename)
                                    attachment_info.append(f"{file.original_filename} -> 附件/{new_filename}")
                                    app.logger.info(f"📁 复制附件: {file.original_filename} -> {new_filename}")
                                else:
                                    app.logger.warning(f"⚠️ 附件不存在: {source_path}")
                                    attachment_info.append(f"{file.original_filename} (文件丢失)")
                        
                        # 将附件信息合并为一个字符串
                        attachment_text = '; '.join(attachment_info) if attachment_info else '无附件'
                        ws.cell(row=row_num, column=col, value=attachment_text)
                        col += 1
                    
                    if row_num % 100 == 1:  # 每100条记录记录一次日志
                        app.logger.info(f"📝 已处理 {row_num - 1} 条Excel记录")
                
                except Exception as row_error:
                    app.logger.error(f"❌ 处理Excel第 {row_num-1} 条记录时出错: {str(row_error)}", exc_info=True)
                    # 继续处理下一条记录
                    continue
            
            # 调整列宽
            for col in range(1, len(headers) + 1):
                max_length = 0
                column = ws.cell(row=1, column=col).column_letter
                for row_cells in ws[column]:
                    try:
                        if len(str(row_cells.value)) > max_length:
                            max_length = len(str(row_cells.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 最大宽度50
                ws.column_dimensions[column].width = max(adjusted_width, 10)  # 最小宽度10
            
            # 保存Excel文件
            excel_filename = f'{file_name}_{timestamp}.xlsx'
            excel_path = os.path.join(temp_dir, excel_filename)
            wb.save(excel_path)
            app.logger.info(f"✅ Excel文件保存完成: {excel_filename}")
            
            # 创建说明文件
            readme_content = f"""导出说明
        ===================

        表单名称: {form_obj.title}
        导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        提交记录数: {len(submissions)}
        附件数量: {len(attachment_files)}

        文件说明:
        - {excel_filename}: 表单提交数据（Excel格式）
        - 附件/: 包含所有提交的附件，文件名格式为"提交ID_原始文件名"

        注意事项:
        - 请维护文件完整性，不要随意修改文件名
        - 附件文件名中的提交ID与Excel表中的提交ID对应
        """
            readme_path = os.path.join(temp_dir, '说明.txt')
            with open(readme_path, 'w', encoding='utf-8') as f:
                f.write(readme_content)
            
            # 创建ZIP文件
            zip_filename = f'{file_name}_{timestamp}.zip'
            zip_path = os.path.join(upload_dir, zip_filename)
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # 添加Excel文件
                zipf.write(excel_path, excel_filename)
                # 添加说明文件
                zipf.write(readme_path, '说明.txt')
                # 添加附件目录
                for root, dirs, files in os.walk(attachments_dir):
                    for file in files:
                        file_path = os.path.join(root, file)
                        arc_name = os.path.join('附件', file)
                        zipf.write(file_path, arc_name)
                
                app.logger.info(f"📁 ZIP文件创建完成: {zip_filename}")
            
            # 清理临时文件
            shutil.rmtree(temp_dir)
            app.logger.info(f"🧹 临时文件已清理: {temp_dir}")
            
            # 返回ZIP文件
            with open(zip_path, 'rb') as f:
                zip_data = f.read()
            
            # 删除临时ZIP文件
            os.remove(zip_path)
            
            response = make_response(zip_data)
            response.headers['Content-Type'] = 'application/zip'
            response.headers['Content-Disposition'] = encode_filename_for_http(zip_filename)
            
            app.logger.info(f"✅ ZIP导出完成，文件名: {zip_filename}, 大小: {len(zip_data)} bytes")
            return response
            
        except Exception as zip_error:
            app.logger.error(f"❌ ZIP导出失败: {str(zip_error)}", exc_info=True)
            # 清理临时文件
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
            if 'zip_path' in locals() and os.path.exists(zip_path):
                os.remove(zip_path)
            raise zip_error
    
    @app.route('/admin/forms/<int:form_id>/delete', methods=['DELETE', 'POST'])
    @login_required
    def admin_delete_form(form_id):
        """删除表单"""
        if not current_user.get_id().startswith('admin_'):
            return jsonify({'error': '无权访问'}), 403
        
        try:
            form_obj = Form.query.get_or_404(form_id)
            
            # 删除相关的上传文件
            submissions = Submission.query.filter_by(form_id=form_id).all()
            for submission in submissions:
                for upload_file in submission.files:
                    file_path = os.path.join(upload_dir, upload_file.saved_filename)
                    if os.path.exists(file_path):
                        try:
                            os.remove(file_path)
                        except Exception as e:
                            app.logger.warning(f"删除文件失败 {file_path}: {str(e)}")
            
            # 删除表单（由于设置了cascade，相关的字段、提交记录、提交数据和上传文件记录会自动删除）
            db.session.delete(form_obj)
            db.session.commit()
            
            if request.is_json:
                return jsonify({'success': True, 'message': '表单删除成功'})
            else:
                flash('表单删除成功', 'success')
                return redirect(url_for('admin_forms'))
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"表单删除失败: {str(e)}")
            if request.is_json:
                return jsonify({'error': '删除失败'}), 500
            else:
                flash('表单删除失败', 'danger')
                return redirect(url_for('admin_forms'))
    
    @app.route('/admin/submissions/<int:submission_id>/update-status', methods=['POST'])
    @login_required
    def admin_update_submission_status(submission_id):
        """更新提交状态"""
        if not current_user.get_id().startswith('admin_'):
            return jsonify({'error': '无权访问'}), 403
        
        try:
            submission = Submission.query.get_or_404(submission_id)
            
            # 获取新状态
            data = request.get_json() if request.is_json else request.form
            new_status = data.get('status')
            
            if new_status not in ['submitted', 'approved', 'rejected']:
                return jsonify({'error': '无效的状态值'}), 400
            
            # 更新状态
            old_status = submission.status
            submission.status = new_status
            db.session.commit()
            
            # 返回成功响应
            status_text = {
                'submitted': '待审核',
                'approved': '已通过', 
                'rejected': '已拒绝'
            }
            
            return jsonify({
                'success': True, 
                'message': f'状态已从"{status_text.get(old_status, old_status)}"更新为"{status_text.get(new_status)}"',
                'new_status': new_status,
                'status_text': status_text.get(new_status)
            })
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"更新提交状态失败: {str(e)}")
            return jsonify({'error': '状态更新失败'}), 500
    
    @app.route('/admin/users/<int:user_id>/toggle-status', methods=['POST'])
    @login_required
    def admin_toggle_user_status(user_id):
        """切换用户状态"""
        if not current_user.get_id().startswith('admin_'):
            return jsonify({'error': '无权访问'}), 403
        
        try:
            user = User.query.get_or_404(user_id)
            
            # 切换状态
            user.is_active = not user.is_active
            db.session.commit()
            
            return jsonify({
                'success': True,
                'is_active': user.is_active,
                'message': f'用户 {user.name} 已{"\u542f\u7528" if user.is_active else "\u7981\u7528"}'
            })
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"用户状态切换失败: {str(e)}")
            return jsonify({'error': '状态更新失败'}), 500
    
    @app.route('/admin/users/batch-action', methods=['POST'])
    @login_required
    def admin_batch_user_action():
        """批量用户操作"""
        if not current_user.get_id().startswith('admin_'):
            return jsonify({'error': '无权访问'}), 403
        
        try:
            data = request.get_json() if request.is_json else request.form
            action = data.get('action')
            user_ids = data.get('user_ids', [])
            
            if not user_ids:
                return jsonify({'error': '请选择要操作的用户'}), 400
            
            if action == 'activate':
                # 批量启用用户
                User.query.filter(User.id.in_(user_ids)).update({'is_active': True}, synchronize_session=False)
                message = f'已成功启用 {len(user_ids)} 个用户'
            elif action == 'deactivate':
                # 批量禁用用户
                User.query.filter(User.id.in_(user_ids)).update({'is_active': False}, synchronize_session=False)
                message = f'已成功禁用 {len(user_ids)} 个用户'
            elif action == 'delete':
                # 批量删除用户（注意：这是危险操作）
                users = User.query.filter(User.id.in_(user_ids)).all()
                for user in users:
                    # 删除用户的所有提交记录和相关文件
                    for submission in user.submissions:
                        for upload_file in submission.files:
                            file_path = os.path.join(upload_dir, upload_file.saved_filename)
                            if os.path.exists(file_path):
                                try:
                                    os.remove(file_path)
                                except Exception as e:
                                    app.logger.warning(f"删除文件失败 {file_path}: {str(e)}")
                    db.session.delete(user)
                message = f'已成功删除 {len(user_ids)} 个用户'
            else:
                return jsonify({'error': '无效的操作类型'}), 400
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': message,
                'action': action,
                'affected_count': len(user_ids)
            })
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"批量用户操作失败: {str(e)}")
            return jsonify({'error': '批量操作失败'}), 500
    
    @app.route('/uploads/<filename>')
    @login_required
    def uploaded_file(filename):
        return send_from_directory(upload_dir, filename)
    
    # API路由
    @app.route('/api/user/profile', methods=['POST'])
    @login_required
    def api_update_user_profile():
        """更新用户个人资料API"""
        if current_user.get_id().startswith('admin_'):
            return jsonify({'error': '管理员不能使用此功能'}), 403
        
        try:
            data = request.get_json() if request.is_json else request.form
            
            # 获取表单数据
            name = data.get('name', '').strip()
            email = data.get('email', '').strip()
            phone = data.get('phone', '').strip()
            current_password = data.get('current_password', '').strip()
            new_password = data.get('new_password', '').strip()
            confirm_password = data.get('confirm_password', '').strip()
            
            # 验证必填字段
            if not name:
                return jsonify({'error': '姓名不能为空'}), 400
            
            # 验证邮箱格式
            if email:
                import re
                email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                if not re.match(email_pattern, email):
                    return jsonify({'error': '邮箱格式不正确'}), 400
                
                # 检查邮箱是否被其他用户使用
                existing_user = User.query.filter(User.email == email, User.id != current_user.id).first()
                if existing_user:
                    return jsonify({'error': '该邮箱已被其他用户使用'}), 400
            
            # 验证手机号
            if phone:
                # 简单的手机号格式验证
                phone_pattern = r'^1[3-9]\d{9}$'
                if not re.match(phone_pattern, phone):
                    return jsonify({'error': '手机号格式不正确'}), 400
                
                # 检查手机号是否被其他用户使用
                existing_user = User.query.filter(User.phone == phone, User.id != current_user.id).first()
                if existing_user:
                    return jsonify({'error': '该手机号已被其他用户使用'}), 400
            
            # 如果要修改密码，验证当前密码
            if new_password:
                if not current_password:
                    return jsonify({'error': '请输入当前密码'}), 400
                
                if not current_user.check_password(current_password):
                    return jsonify({'error': '当前密码错误'}), 400
                
                if new_password != confirm_password:
                    return jsonify({'error': '新密码和确认密码不一致'}), 400
                
                if len(new_password) < 6:
                    return jsonify({'error': '密码长度不能少于6位'}), 400
            
            # 更新用户信息
            current_user.name = name
            current_user.email = email if email else None
            current_user.phone = phone if phone else None
            
            # 如果要修改密码
            if new_password:
                current_user.set_password(new_password)
            
            db.session.commit()
            
            app.logger.info(f"👤 用户 {current_user.id} 更新个人资料成功")
            
            return jsonify({
                'success': True,
                'message': '个人信息更新成功！',
                'user': {
                    'name': current_user.name,
                    'email': current_user.email,
                    'phone': current_user.phone
                }
            })
            
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"❌ 更新用户资料失败: {str(e)}")
            return jsonify({'error': f'更新失败: {str(e)}'}), 500
    
    @app.route('/api/forms/<int:form_id>/toggle')
    @login_required
    def api_toggle_form(form_id):
        if not current_user.get_id().startswith('admin_'):
            return jsonify({'error': '无权访问'}), 403
        
        form_obj = Form.query.get_or_404(form_id)
        form_obj.is_active = not form_obj.is_active
        db.session.commit()
        
        return jsonify({'success': True, 'is_active': form_obj.is_active})
    
    # 初始化数据库(在应用上下文中完成，由run.py调用)
    
    return app

        if __name__ == '__main__':
    app = create_app()
    app.run(debug=True, host='0.0.0.0', port=5000)
